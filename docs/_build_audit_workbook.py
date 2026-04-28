"""
Build big12-dept-workflow-audit.xlsx per the spec in
~/.claude/plans/delegated-puzzling-church.md (Department Workflow Audit Spreadsheet section).

One-shot document-creation tool. Produces an empty template with 13 tabs
(Index, Scoring Rollup, 11 department tabs). Nick populates it by running
the /dept-ai-audit skill per department.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

OUT = Path.home() / "HELiiX" / "docs" / "big12-dept-workflow-audit.xlsx"

# ---------- Department tabs ----------
DEPARTMENTS = [
    "Office of the Commissioner",
    "Communications",
    "Compliance & Governance",
    "Content & Brand Marketing",
    "Partnerships & Ticketing",
    "Competition & Events",
    "Legal",
    "Member Admin & Office Tech",
    "Academics & Student Services",
    "Board & Meetings",
    "HR & Finance",
]

# ---------- Section A fields (Department Profile) ----------
PROFILE_FIELDS = [
    "Mission / core responsibilities",
    "Internal customers",
    "External stakeholders",
    "Recurring outputs",
    "Key deadlines / event cycles / seasonal peaks",
    "Tools & systems currently used",
    "Data sources relied on",
    "Approval chains",
    "Regulatory / legal / privacy constraints",
    "Failure modes if work is done poorly",
    "How success is measured",
]

# ---------- Section B columns (Role Map) ----------
ROLE_MAP_COLS = [
    "Title",
    "Current holder (metadata)",
    "Responsibilities",
    "Recurring tasks & docs created",
    "Systems touched",
    "Decisions owned",
    "Approvals needed",
    "Pain points",
    "What good work looks like",
    "Evidence tag",
]

# ---------- Section C + D columns (Workflow Inventory + Scoring Matrix) ----------
WORKFLOW_COLS = [
    # Section C — Workflow Inventory
    "Workflow name",                      # A
    "Owner (role)",                       # B
    "Trigger",                            # C
    "Cadence",                            # D
    "Inputs",                             # E
    "Systems used",                       # F
    "Human steps (numbered)",             # G
    "Outputs",                            # H
    "Downstream consumers",               # I
    "Approval gates",                     # J
    "Manual effort (hrs / occurrence)",   # K
    "Pain points",                        # L
    "Risk if wrong (L/M/H/Critical)",     # M   -- dropdown
    "Evidence tag",                       # N   -- dropdown
    "Tags (classification)",              # O
    # Section D — Scoring Matrix
    "Pain (1-5)",                         # P   -- dropdown
    "Frequency (1-5)",                    # Q   -- dropdown
    "Risk if Wrong (1-5)",                # R   -- dropdown
    "AI Fit (1-5)",                       # S   -- dropdown
    "Data Readiness (1-5)",               # T   -- dropdown
    "Change Mgmt Difficulty (1-5)",       # U   -- dropdown
    "Total score",                        # V   -- formula
    "Pain gate",                          # W   -- dropdown Pass/Fail
    "Data gate",                          # X   -- dropdown Pass/Fail
    "Safety gate",                        # Y   -- dropdown Pass/Fail
    "Advances to Step 2?",                # Z   -- dropdown Y/N
]

# ---------- Styles ----------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=13)
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
SUB_HEADER_FONT = Font(bold=True, size=11)
SUB_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
FIELD_LABEL_FONT = Font(bold=True, size=11)
FIELD_LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_section_header(ws, row, text, last_col_letter):
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    c = ws[f"A{row}"]
    c.value = text
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def write_field_label_col(ws, row, label):
    c = ws.cell(row=row, column=1, value=label)
    c.font = FIELD_LABEL_FONT
    c.fill = FIELD_LABEL_FILL
    c.alignment = WRAP_ALIGN
    c.border = BORDER
    for col in range(2, 5):
        b = ws.cell(row=row, column=col)
        b.alignment = WRAP_ALIGN
        b.border = BORDER


def write_column_headers(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 32


def add_validation(ws, col_letter, values, first_row, last_row):
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Invalid value"
    dv.errorTitle = "Must be one of: " + ", ".join(values)
    dv.prompt = "Pick one: " + ", ".join(values)
    dv.promptTitle = "Allowed values"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def build_department_tab(ws):
    # ---- Section A: Department Profile (rows 1-15) ----
    set_section_header(ws, 1, "Section A  —  Department Profile", "D")
    # Sub-header row
    sub = ["Field", "Value", "Evidence tag", "Notes"]
    for i, h in enumerate(sub, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = SUB_HEADER_FONT
        c.fill = SUB_HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = BORDER
    ws.row_dimensions[2].height = 20

    for i, field in enumerate(PROFILE_FIELDS):
        r = 3 + i
        write_field_label_col(ws, r, field)
        ws.row_dimensions[r].height = 30

    # Evidence-tag dropdown for profile rows (col C, rows 3..13)
    add_validation(ws, "C", ["Confirmed", "Inferred", "Speculative"], 3, 13)

    # ---- Section B: Role Map (rows 17-30) ----
    set_section_header(ws, 17, "Section B  —  Role Map", "J")
    write_column_headers(ws, 18, ROLE_MAP_COLS)
    # blank role rows 19-30 (12 role rows), styled with borders
    for r in range(19, 31):
        for col in range(1, len(ROLE_MAP_COLS) + 1):
            c = ws.cell(row=r, column=col)
            c.alignment = WRAP_ALIGN
            c.border = BORDER
        ws.row_dimensions[r].height = 28
    # Evidence-tag dropdown on Section B (col J, rows 19-30)
    add_validation(ws, "J", ["Confirmed", "Inferred", "Speculative"], 19, 30)

    # ---- Section C + D: Workflow Inventory + Scoring Matrix (rows 32+) ----
    set_section_header(ws, 32, "Section C + D  —  Workflow Inventory and Scoring Matrix", "Z")
    write_column_headers(ws, 33, WORKFLOW_COLS)

    WF_FIRST = 34
    WF_LAST = 73  # 40 workflow rows

    # Style workflow rows + add Total-score formula
    for r in range(WF_FIRST, WF_LAST + 1):
        for col in range(1, len(WORKFLOW_COLS) + 1):
            c = ws.cell(row=r, column=col)
            c.alignment = WRAP_ALIGN
            c.border = BORDER
        # Total score formula in column V (col 22) = SUM(P:U)
        ws.cell(
            row=r, column=22,
            value=f"=IFERROR(SUM(P{r}:U{r}),\"\")"
        )
        ws.row_dimensions[r].height = 30

    # Dropdowns on Section C + D columns
    # M: Risk if wrong
    add_validation(ws, "M", ["Low", "Medium", "High", "Critical"], WF_FIRST, WF_LAST)
    # N: Evidence tag
    add_validation(ws, "N", ["Confirmed", "Inferred", "Speculative"], WF_FIRST, WF_LAST)
    # P-U: scoring 1-5
    for col_letter in ["P", "Q", "R", "S", "T", "U"]:
        add_validation(ws, col_letter, ["1", "2", "3", "4", "5"], WF_FIRST, WF_LAST)
    # W-Y: Pass/Fail gates
    for col_letter in ["W", "X", "Y"]:
        add_validation(ws, col_letter, ["Pass", "Fail"], WF_FIRST, WF_LAST)
    # Z: Y/N advances
    add_validation(ws, "Z", ["Y", "N"], WF_FIRST, WF_LAST)

    # Freeze panes so header band stays visible
    ws.freeze_panes = "A34"

    # Column widths
    widths = {
        "A": 28, "B": 22, "C": 22, "D": 16, "E": 28, "F": 22,
        "G": 32, "H": 24, "I": 22, "J": 22, "K": 18, "L": 28,
        "M": 18, "N": 16, "O": 28,
        "P": 10, "Q": 12, "R": 14, "S": 10, "T": 14, "U": 18,
        "V": 12, "W": 12, "X": 12, "Y": 12, "Z": 16,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    # Section A column widths
    ws.column_dimensions["A"].width = 34  # field labels
    ws.column_dimensions["B"].width = 60  # value
    ws.column_dimensions["C"].width = 16  # evidence tag
    ws.column_dimensions["D"].width = 40  # notes

    # Note: workflow inventory starts below Section A/B; column A needs
    # to accommodate BOTH the profile field labels (wide) and workflow names.
    # Width 34 works for both.


def build_index_tab(ws, department_names):
    ws.title = "Index"
    set_section_header(ws, 1, "Big 12 Department Workflow Audit  —  Index", "F")
    headers = [
        "Department",
        "Tab link",
        "# workflows inventoried",
        "# in top 5 (advances to Step 2)",
        "Pilot candidate? (Y/N)",
        "Owner",
    ]
    write_column_headers(ws, 2, headers)

    for i, dept in enumerate(department_names):
        r = 3 + i
        ws.cell(row=r, column=1, value=dept).alignment = WRAP_ALIGN
        # Hyperlink to that department's tab (A1)
        link_cell = ws.cell(row=r, column=2, value=f"Go to {dept}")
        link_cell.hyperlink = f"#'{dept}'!A1"
        link_cell.font = Font(color="0563C1", underline="single")
        # Formula counting non-empty workflow names in that dept's Section C
        ws.cell(
            row=r, column=3,
            value=f"=COUNTA('{dept}'!A34:A73)"
        )
        # Count of workflows with "Y" in column Z (advances to Step 2)
        ws.cell(
            row=r, column=4,
            value=f'=COUNTIF(\'{dept}\'!Z34:Z73,"Y")'
        )
        # Pilot candidate column (dropdown)
        # Owner column left blank for Nick to fill
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=r, column=col)
            c.alignment = WRAP_ALIGN
            c.border = BORDER
        ws.row_dimensions[r].height = 24

    # Dropdown on pilot-candidate column (col 5, rows 3..13)
    add_validation(ws, "E", ["Y", "N"], 3, 2 + len(department_names))

    # Column widths
    widths = [32, 22, 28, 36, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


def build_rollup_tab(ws, department_names):
    ws.title = "Scoring Rollup"
    set_section_header(ws, 1, "Cross-Department Scoring Rollup", "F")
    ws.cell(row=2, column=1, value=(
        "This tab rolls up workflows that advance to Step 2 across all "
        "departments. Populate Section D (scoring) in each department tab; "
        "the rollup surfaces the conference-wide top candidates."
    )).alignment = WRAP_ALIGN
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 40

    # Headers
    headers = [
        "Department",
        "Workflow name",
        "Owner (role)",
        "Total score",
        "Risk if wrong",
        "Advances to Step 2?",
    ]
    write_column_headers(ws, 4, headers)

    # For each department, pull workflow rows that advance to Step 2.
    # Openpyxl can't easily do a true cross-sheet query, so we pre-populate
    # per-department blocks with pass-through formulas for the first 10 rows
    # of each department. Nick can adjust / extend as needed.
    current_row = 5
    for dept in department_names:
        # Section header for this department within the rollup
        ws.merge_cells(f"A{current_row}:F{current_row}")
        hc = ws.cell(row=current_row, column=1, value=dept)
        hc.font = SUB_HEADER_FONT
        hc.fill = SUB_HEADER_FILL
        hc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[current_row].height = 20
        current_row += 1
        # Pass-through formula rows for first 10 workflows of that dept
        for i in range(10):
            src_row = 34 + i
            ws.cell(row=current_row, column=1, value=dept).alignment = WRAP_ALIGN
            ws.cell(
                row=current_row, column=2,
                value=f"=IFERROR('{dept}'!A{src_row},\"\")"
            )
            ws.cell(
                row=current_row, column=3,
                value=f"=IFERROR('{dept}'!B{src_row},\"\")"
            )
            ws.cell(
                row=current_row, column=4,
                value=f"=IFERROR('{dept}'!V{src_row},\"\")"
            )
            ws.cell(
                row=current_row, column=5,
                value=f"=IFERROR('{dept}'!M{src_row},\"\")"
            )
            ws.cell(
                row=current_row, column=6,
                value=f"=IFERROR('{dept}'!Z{src_row},\"\")"
            )
            for col in range(1, 7):
                c = ws.cell(row=current_row, column=col)
                c.alignment = WRAP_ALIGN
                c.border = BORDER
            current_row += 1

    widths = [32, 36, 24, 14, 18, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"


def main():
    wb = Workbook()

    # First sheet is created by default -> rename to Index
    idx_sheet = wb.active
    build_index_tab(idx_sheet, DEPARTMENTS)

    # Second sheet: Scoring Rollup
    rollup_sheet = wb.create_sheet("Scoring Rollup")
    build_rollup_tab(rollup_sheet, DEPARTMENTS)

    # One sheet per department
    for dept in DEPARTMENTS:
        ws = wb.create_sheet(dept)
        build_department_tab(ws)

    wb.save(OUT)
    print(f"wrote {OUT}  ({wb.sheetnames[:3]}... total {len(wb.sheetnames)} sheets)")


if __name__ == "__main__":
    main()
